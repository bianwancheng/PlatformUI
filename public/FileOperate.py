#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2019/11/19 11:32 
# @Author : Wancheng.b 
# @File : FileOperate.py 
# @Software: PyCharm

'''
用于接收UI传过来的post表单生成一个yaml文件或者excel，方便调试
然后读取文件进行操作

'''

from openpyxl import load_workbook


class FileOperate:
    '''
    读取excel返回列表，列表元素为字典
    '''

    def read_excel(self):
        wb = load_workbook('D:\pycharmwork\PlatformUI\TestCase\login\ModelTest.xlsx')
        sheetnames = wb.sheetnames  # 获取所有sheetname
        core_list = []
        for k in range(0, len(sheetnames)):
            sheet = wb[sheetnames[k]]
            rows = sheet.max_row  # 获取该sheet中行数最大的值
            columns = sheet.max_column  # 获取该sheet中列数最大的值
            row_1 = []
            # print(rows, columns)
            # 遍历sheet中的值
            for i in range(1, rows + 1):
                dict = {}
                for j in range(1, columns + 1):
                    value = sheet.cell(i, j).value
                    if i == 1:
                        row_1.append(value)
                    else:
                        dict[row_1[j - 1]] = value
                core_list.append(dict)
        return core_list[1:]


if __name__ == '__main__':
    print(FileOperate().read_excel())
